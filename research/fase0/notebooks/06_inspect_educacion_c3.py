"""Ver columnas completas de educacion_c3 (nivel educativo alcanzado)."""
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

path = Path(__file__).resolve().parents[1] / "data" / "censo" / "c2022_caba_educacion_c3_1.xlsx"
wb = load_workbook(path, read_only=True, data_only=True)
ws = wb["Cuadro 3.1.1"]  # Comuna 1
print(f"Dimensiones: {ws.max_row} x {ws.max_column}")

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i >= 12:
        break
    cells = []
    for j, v in enumerate(row):
        s = "" if v is None else str(v).replace("\xa0", " ").replace("\ufffd", "_")[:22]
        cells.append(f"{j}:{s}")
    print(f"[{i:2d}] {' | '.join(cells)}")

wb.close()
