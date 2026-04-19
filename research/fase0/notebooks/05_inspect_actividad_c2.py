"""
Inspeccion detallada de actividad_economica_c2_1.xlsx - hoja Comuna 1 completa.
Queremos entender: estructura exacta de headers, posicion de totales, que
columnas representan cada condicion de actividad.
"""
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

path = Path(__file__).resolve().parents[1] / "data" / "censo" / "c2022_caba_actividad_economica_c2_1.xlsx"
wb = load_workbook(path, read_only=True, data_only=True)

# Total CABA primero
sheet_total = [s for s in wb.sheetnames if "Cuadro" in s and s.replace("Cuadro", "").strip() == "2.1"]
print("Hojas:", wb.sheetnames[:5], "...")
print(f"Hoja total candidata: {sheet_total}")

# Vamos por Cuadro 2.1.1 = Comuna 1
target = next((s for s in wb.sheetnames if "2.1.1" == s.split()[-1]), None)
print(f"\nInspeccionando hoja: '{target}'")

ws = wb[target]
print(f"Dimensiones: {ws.max_row} x {ws.max_column}\n")

for i, row in enumerate(ws.iter_rows(values_only=True)):
    if i >= 25:
        break
    cells = []
    for v in row[:20]:
        if v is None:
            cells.append("")
        else:
            s = str(v)
            if any(ord(c) > 127 for c in s):
                # limpiar caracteres raros
                s = s.encode("latin-1", "replace").decode("utf-8", "replace")
                s = s.replace("\ufffd", "").replace("\xa0", " ")
            cells.append(s[:30])
    print(f"[{i:2d}] {' | '.join(cells)}")

wb.close()
