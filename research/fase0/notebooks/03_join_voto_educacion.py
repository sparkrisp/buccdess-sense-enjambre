"""
Fase 0 - JOIN electoral x educacion por comuna.
Objetivo: validar que los datos se pueden cruzar y que la correlacion es plausible.

Hipotesis: comunas con mayor proporcion de universitarios completos votaron diferente
que comunas con menor proporcion. Si sale correlacion clara -> GO para Fase 1.
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CABA = ROOT / "data" / "caba" / "generales_2023_caba.csv"
CENSO = ROOT / "data" / "censo"

# ---------- 1) Resultados electorales por comuna (Jefe de Gobierno) ----------
df = pd.read_csv(CABA, sep=";")
jef = df[(df["cargo_nombre"] == "JEF") & (df["votos_tipo"] == "POSITIVO")]
elect = (
    jef.groupby(["seccion_id", "agrupacion_nombre"])["votos_cantidad"]
    .sum()
    .unstack(fill_value=0)
)
elect["TOTAL_POSITIVOS"] = elect.sum(axis=1)
for col in ["UNION POR LA PATRIA", "JUNTOS POR EL CAMBIO", "LA LIBERTAD AVANZA"]:
    elect[f"pct_{col}"] = (elect[col] / elect["TOTAL_POSITIVOS"] * 100).round(1)
elect = elect.reset_index().rename(columns={"seccion_id": "comuna"})

# ---------- 2) Dump raw del cuadro de educacion Comuna 1 para entender formato ----------
edu_file = CENSO / "c2022_caba_educacion_c3_1.xlsx"
# Educacion c3 suele ser: poblacion por nivel educativo alcanzado
wb = load_workbook(edu_file, read_only=True, data_only=True)
print(f"Hojas de {edu_file.name}: {wb.sheetnames}")

# Buscar la hoja del total CABA (sin subindice)
total_sheet = None
for name in wb.sheetnames:
    if "3.1" in name and "." not in name.replace("3.1", "", 1):
        total_sheet = name
        break

# Mostrar primeras 30 filas de la primera hoja de datos (la del total)
data_sheets = [s for s in wb.sheetnames if "Cuadro" in s or "uadro" in s]
print(f"Hojas de datos: {data_sheets[:3]}")
first_data = data_sheets[0] if data_sheets else None

if first_data:
    ws = wb[first_data]
    print(f"\n--- Hoja '{first_data}' ({ws.max_row}x{ws.max_column}) ---")
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        if i >= 35:
            break
        non_empty = [str(v).strip() if v is not None else "" for v in row[:12]]
        print(f"  [{i:2d}] {' | '.join(non_empty)}")

wb.close()

# ---------- 3) Mostrar resultado electoral ya agregado para comparar ----------
print("\n\n=== RESULTADOS JEFE DE GOBIERNO POR COMUNA ===")
print(elect[["comuna", "pct_JUNTOS POR EL CAMBIO", "pct_UNION POR LA PATRIA", "pct_LA LIBERTAD AVANZA", "TOTAL_POSITIVOS"]].to_string(index=False))
