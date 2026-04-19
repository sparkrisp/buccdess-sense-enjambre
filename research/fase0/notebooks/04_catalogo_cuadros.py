"""
Fase 1 - Catalogar cuadros del Censo 2022 CABA.
Para cada archivo xlsx: leer la hoja 'Indice' y extraer el titulo del Cuadro N.1
(total CABA) para saber que variable cubre.
"""
import sys
from pathlib import Path
from openpyxl import load_workbook

sys.stdout.reconfigure(encoding="utf-8")

CENSO = Path(__file__).resolve().parents[1] / "data" / "censo"


def decode_latin(s):
    """Los textos vienen con doble-encoding latin-1->utf-8. Recuperamos utf-8."""
    if s is None:
        return ""
    try:
        # openpyxl lee bytes como si fueran latin-1, pero el contenido original es utf-8.
        # Re-encodeamos a latin-1 (sin perdida) y decodeamos como utf-8.
        return str(s).encode("latin-1", errors="replace").decode("utf-8", errors="replace")
    except Exception:
        return str(s)


def get_first_cuadro_title(path: Path) -> str:
    """Leer la primera linea util de la hoja Indice."""
    try:
        wb = load_workbook(path, read_only=True, data_only=True)
        # La hoja indice puede llamarse 'Indice' o 'Índice' (con encoding roto)
        indice = None
        for sheet in wb.sheetnames:
            if "ndice" in sheet.lower() or "indice" in sheet.lower():
                indice = sheet
                break
        if not indice:
            return "[sin indice]"
        ws = wb[indice]
        for row in ws.iter_rows(values_only=True, max_row=10):
            for cell in row:
                if cell and "Cuadro" in str(cell) and ("." in str(cell)):
                    wb.close()
                    return decode_latin(cell)[:200]
        wb.close()
        return "[no encontrado]"
    except Exception as e:
        return f"[error: {e}]"


priority_prefixes = ["pob_c", "educacion_c", "actividad_economica_c", "hogares_c"]

files_by_group: dict[str, list[Path]] = {p: [] for p in priority_prefixes}
for f in sorted(CENSO.glob("*.xlsx")):
    for p in priority_prefixes:
        if p in f.name:
            files_by_group[p].append(f)
            break

for group, files in files_by_group.items():
    print(f"\n{'='*80}")
    print(f"GRUPO: {group}")
    print(f"{'='*80}")
    for f in files:
        title = get_first_cuadro_title(f)
        print(f"\n>>> {f.name}")
        print(f"    {title}")
