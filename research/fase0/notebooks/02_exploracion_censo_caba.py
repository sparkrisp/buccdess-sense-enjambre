"""
Fase 0 - Inspeccion de tabulados del Censo 2022 CABA.
Objetivo: entender el formato (INDEC suele poner multi-index y celdas combinadas)
y validar que el join con datos electorales por comuna es viable.
"""
import pandas as pd
from pathlib import Path
from openpyxl import load_workbook

DATA = Path(__file__).resolve().parents[1] / "data" / "censo"

def peek(fname: str, rows: int = 15):
    path = DATA / fname
    print("\n" + "=" * 70)
    print(f"ARCHIVO: {fname}")
    print("=" * 70)
    wb = load_workbook(path, read_only=True, data_only=True)
    print(f"Hojas: {wb.sheetnames}")
    for sheet_name in wb.sheetnames[:2]:
        ws = wb[sheet_name]
        print(f"\n--- Hoja: '{sheet_name}' ({ws.max_row} filas x {ws.max_column} cols) ---")
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= rows:
                break
            non_empty = [str(v) if v is not None else "" for v in row]
            print(f"  [{i:2d}] {' | '.join(non_empty[:8])}")
    wb.close()

for f in [
    "c2022_caba_pob_c1_1.xlsx",
    "c2022_caba_pob_c4_1.xlsx",
    "c2022_caba_educacion_c1_1.xlsx",
    "c2022_caba_actividad_economica_c1_1.xlsx",
    "c2022_caba_hogares_c1_1.xlsx",
]:
    peek(f, rows=18)
