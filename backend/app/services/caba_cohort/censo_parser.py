"""
Parser de tabulados del Censo 2022 CABA (INDEC).

Los xlsx tienen:
- Una hoja "Cuadro N.1" con el total CABA
- 15 hojas "Cuadro N.1.K" con la desagregacion por Comuna K (1..15)
- Header multi-nivel en filas 2-4
- Bloques por sexo: "Total" (fila ~5), "Mujer/Femenino" (~21), "Varon/Masculino" (~38)
- Filas de edad quinquenal: "14", "15-19", "20-24", ..., "80 y mas"
"""
from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


# Grupos etarios que aparecen en los cuadros INDEC (14+ anos)
AGE_GROUPS_14PLUS = [
    "14", "15-19", "20-24", "25-29", "30-34", "35-39",
    "40-44", "45-49", "50-54", "55-59", "60-64",
    "65-69", "70-74", "75-79", "80+",
]


def _clean(v) -> str:
    """Normalizar strings con encoding roto y NBSP."""
    if v is None:
        return ""
    s = str(v).strip()
    return s.replace("\xa0", " ").replace("\u00a0", " ")


def _norm_age(s: str) -> str | None:
    """Normalizar etiqueta de grupo etario. '80 y mas' -> '80+', etc."""
    s = _clean(s).lower()
    if not s:
        return None
    if "y m" in s or "80" in s and "+" not in s and "-" not in s:
        return "80+" if "80" in s else None
    for g in AGE_GROUPS_14PLUS:
        if g == s:
            return g
    # Solo el numero 14 aislado
    if s == "14":
        return "14"
    # Rangos tipo "15-19"
    if "-" in s and s.replace("-", "").isdigit():
        return s
    return None


def _norm_sex(s: str) -> str | None:
    s = _clean(s).lower()
    if "mujer" in s or "femenino" in s:
        return "F"
    if "var" in s or "masculino" in s:
        return "M"
    if s == "total":
        return "T"
    return None


@dataclass
class ActividadRow:
    comuna: int
    sexo: str  # 'F' | 'M'
    edad: str  # AGE_GROUPS_14PLUS
    poblacion_14plus: int
    pea_total: int  # economicamente activa
    ocupada: int
    desocupada: int
    pnea: int  # no economicamente activa


def parse_actividad_comuna(path: Path, comuna: int) -> list[ActividadRow]:
    """
    Parsea la hoja 'Cuadro 2.1.{comuna}' de actividad_economica_c2_1.xlsx.
    Layout de columnas (0-indexed):
      0: sexo (en filas de bloque) / vacio en filas de edad
      1: edad
      2: Poblacion 14+ total
      3: PEA Total
      4: Ocupada
      5: Desocupada
      6: PNEA
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = f"Cuadro 2.1.{comuna}"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No existe hoja '{sheet_name}' en {path.name}")
    ws = wb[sheet_name]

    current_sex: str | None = None
    rows: list[ActividadRow] = []

    for row in ws.iter_rows(values_only=True):
        col0 = _clean(row[0]) if row[0] is not None else ""
        col1 = _clean(row[1]) if len(row) > 1 and row[1] is not None else ""

        # Fila con nombre de sexo (Total / Mujer / Varon)
        sex_candidate = _norm_sex(col0) if col0 else None
        if sex_candidate:
            current_sex = sex_candidate
            # Puede o no tener datos en la misma fila (Total agrega todas las edades,
            # no nos interesa porque lo podemos reconstruir)
            continue

        # Fila con edad
        if current_sex in ("F", "M"):
            age = _norm_age(col1)
            if age and age in AGE_GROUPS_14PLUS:
                try:
                    rows.append(
                        ActividadRow(
                            comuna=comuna,
                            sexo=current_sex,
                            edad=age,
                            poblacion_14plus=int(row[2] or 0),
                            pea_total=int(row[3] or 0),
                            ocupada=int(row[4] or 0),
                            desocupada=int(row[5] or 0),
                            pnea=int(row[6] or 0),
                        )
                    )
                except (TypeError, ValueError):
                    pass

    wb.close()
    return rows


def load_actividad_all_comunas(path: Path) -> pd.DataFrame:
    """DataFrame tidy para las 15 comunas CABA."""
    all_rows: list[ActividadRow] = []
    for comuna in range(1, 16):
        all_rows.extend(parse_actividad_comuna(path, comuna))
    return pd.DataFrame([r.__dict__ for r in all_rows])


# ---------------------------------------------------------------------------
# Educacion: nivel educativo maximo alcanzado (educacion_c3_1.xlsx)
# ---------------------------------------------------------------------------

EDU_LEVELS = ["ninguno", "primario", "secundario", "terciario", "universitario"]


def _to_int(v) -> int:
    """Convertir celda a int. '///' y '-' son 0. None tambien."""
    if v is None:
        return 0
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip()
    if s in ("", "-", "///"):
        return 0
    try:
        return int(s)
    except ValueError:
        return 0


@dataclass
class EducacionRow:
    comuna: int
    sexo: str  # 'F' | 'M'
    edad: str  # grupos quinquenales del censo (desde 15-19 en adelante + '14')
    poblacion_5plus: int  # Total poblacion 5+ en ese grupo
    ninguno: int         # sin instruccion + primario incompleto + EGB incompleto
    primario: int        # primario completo + EGB completo
    secundario: int      # secundario (inc+comp) + polimodal (inc+comp)
    terciario: int       # terciario (inc+comp) + universitario incompleto
    universitario: int   # universitario completo + posgrado (inc+comp)


def parse_educacion_comuna(path: Path, comuna: int) -> list[EducacionRow]:
    """
    Parsea la hoja 'Cuadro 3.1.{comuna}' de educacion_c3_1.xlsx.
    Layout (0-indexed):
      0: sexo (Total/Mujer.../Varon...) en filas de bloque
      1: edad (simple o quinquenal)
      2: Poblacion 5+ total
      3: Poblacion 5+ que asistio
      4: Sin instruccion
      5,6,7: Primario (Total,Inc,Comp)
      8,9,10: EGB (Total,Inc,Comp)
      11,12,13: Secundario (Total,Inc,Comp)
      14,15,16: Polimodal (Total,Inc,Comp)
      17,18,19: Terciario (Total,Inc,Comp)
      20,21,22: Universitario (Total,Inc,Comp)
      23,24,25: Posgrado (Total,Inc,Comp)
      26: Ignorado
    """
    wb = load_workbook(path, read_only=True, data_only=True)
    sheet_name = f"Cuadro 3.1.{comuna}"
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"No existe hoja '{sheet_name}' en {path.name}")
    ws = wb[sheet_name]

    current_sex: str | None = None
    rows: list[EducacionRow] = []

    for row in ws.iter_rows(values_only=True):
        col0 = _clean(row[0]) if row[0] is not None else ""
        col1 = _clean(row[1]) if len(row) > 1 and row[1] is not None else ""

        sex_candidate = _norm_sex(col0)
        if sex_candidate:
            current_sex = sex_candidate
            continue

        if current_sex not in ("F", "M"):
            continue

        age = _norm_age(col1)
        if not age or age not in AGE_GROUPS_14PLUS:
            continue

        poblacion_5plus = _to_int(row[2])
        sin_instruccion = _to_int(row[4])
        prim_inc = _to_int(row[6])
        prim_comp = _to_int(row[7])
        egb_inc = _to_int(row[9])
        egb_comp = _to_int(row[10])
        sec_total = _to_int(row[11])
        poli_total = _to_int(row[14])
        terc_total = _to_int(row[17])
        uni_inc = _to_int(row[21])
        uni_comp = _to_int(row[22])
        posg_total = _to_int(row[23])

        rows.append(
            EducacionRow(
                comuna=comuna,
                sexo=current_sex,
                edad=age,
                poblacion_5plus=poblacion_5plus,
                ninguno=sin_instruccion + prim_inc + egb_inc,
                primario=prim_comp + egb_comp,
                secundario=sec_total + poli_total,
                terciario=terc_total + uni_inc,
                universitario=uni_comp + posg_total,
            )
        )

    wb.close()
    return rows


def load_educacion_all_comunas(path: Path) -> pd.DataFrame:
    all_rows: list[EducacionRow] = []
    for comuna in range(1, 16):
        all_rows.extend(parse_educacion_comuna(path, comuna))
    return pd.DataFrame([r.__dict__ for r in all_rows])


# ---------------------------------------------------------------------------
# Master: combinar actividad + educacion
# ---------------------------------------------------------------------------

def load_censo_master(censo_dir: Path) -> pd.DataFrame:
    """
    DataFrame master: 15 comunas x 2 sexos x 15 grupos etarios, con columnas
    de actividad (ocupada/desocupada/pnea) y nivel educativo alcanzado.

    Join por (comuna, sexo, edad). Se usa la poblacion 14+ de actividad como
    base para calcular shares educativos (que incluyen 5+ pero aproximamos).
    """
    act = load_actividad_all_comunas(censo_dir / "c2022_caba_actividad_economica_c2_1.xlsx")
    edu = load_educacion_all_comunas(censo_dir / "c2022_caba_educacion_c3_1.xlsx")

    df = act.merge(edu, on=["comuna", "sexo", "edad"], how="left", suffixes=("_act", "_edu"))

    # Fillna: el grupo '14' (14 anios solos) esta en actividad pero no en educacion
    # (que tiene '10-14' como quinquenal). Rellenamos con 0 y el sampler lo trata
    # como distribucion educativa uniforme para esa celda.
    edu_cols = ["poblacion_5plus", "ninguno", "primario", "secundario", "terciario", "universitario"]
    for col in edu_cols:
        if col in df.columns:
            df[col] = df[col].fillna(0)
    return df


if __name__ == "__main__":
    # Smoke test
    import sys
    sys.stdout.reconfigure(encoding="utf-8")

    censo_dir = Path(__file__).resolve().parents[1] / "data" / "censo"
    master = load_censo_master(censo_dir)

    print(f"Master shape: {master.shape}")
    print(f"Columnas: {list(master.columns)}")
    print("\nSample (Comuna 1, mujeres):")
    sample = master[(master["comuna"] == 1) & (master["sexo"] == "F")].head(5)
    print(sample.to_string(index=False))

    print("\nTotales por comuna (poblacion 14+):")
    print(master.groupby("comuna")["poblacion_14plus"].sum().to_string())

    print("\nNivel educativo por comuna (% 14+ con universitario completo):")
    master["tot_edu"] = (
        master["ninguno"] + master["primario"] + master["secundario"]
        + master["terciario"] + master["universitario"]
    )
    agg = master.groupby("comuna")[["tot_edu", "universitario"]].sum()
    agg["pct_uni"] = (agg["universitario"] / agg["tot_edu"] * 100).round(1)
    print(agg.to_string())
